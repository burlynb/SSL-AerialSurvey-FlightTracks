"""
make_track_map.py
-----------------
Generates index.html with interactive satellite map of SSL aerial survey flight
tracks.

Survey data:
  Gulf of Alaska (GOA): 2021 xlsx + 2024 csv
  Aleutian Islands (ALI): 2022 xlsx/csv + 2023 xlsx/csv

Photos loaded from:
  photos/2021/  photos/2024/  photos/2022/  photos/2023/

UI: region toggle (Gulf of Alaska | Aleutian Islands),
    year sub-toggle (2021/Both/2024 for GOA; 2022/Both/2023 for ALI).
"""

import csv
import openpyxl
import folium
import re
import math
import glob
from collections import defaultdict
from pathlib import Path

# ── NMEA coordinate conversion ─────────────────────────────────────────────────

def nmea_to_dd(val, hemisphere):
    """Convert NMEA DDDMM.MMMMM to decimal degrees (signed)."""
    v = float(val)
    deg = int(v / 100)
    minutes = v - deg * 100
    dd = deg + minutes / 60.0
    if hemisphere.upper() in ('S', 'W'):
        dd = -dd
    return dd

# ── bearing + arrow helpers ────────────────────────────────────────────────────

def compass_bearing(lat1, lon1, lat2, lon2):
    lat1, lat2 = math.radians(lat1), math.radians(lat2)
    dlon = math.radians(lon2 - lon1)
    x = math.sin(dlon) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(dlon)
    return (math.degrees(math.atan2(x, y)) + 360) % 360

def stable_bearing(coords, pct_from, pct_to):
    n = len(coords)
    i = max(0, int(n * pct_from))
    j = min(n - 1, int(n * pct_to))
    if i == j:
        j = min(n - 1, i + 1)
    return compass_bearing(*coords[i], *coords[j])

def arrowhead_html(bearing_deg, color, size=20):
    h = size
    w = int(size * 0.75)
    css_rot = bearing_deg - 90
    shadow = "drop-shadow(0px 0px 3px rgba(0,0,0,1))"
    return (
        f'<div data-mtype="detail">'
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{w}" height="{h}" '
        f'viewBox="0 0 {w} {h}" '
        f'style="transform:rotate({css_rot:.1f}deg);transform-origin:center center;'
        f'filter:{shadow};overflow:visible;">'
        f'<polygon points="0,0 {w},{h//2} 0,{h}" fill="{color}"/>'
        f'</svg></div>'
    )

def place_arrowhead(location, bearing_deg, color, group, tooltip='', size=20):
    h, w = size, int(size * 0.75)
    folium.Marker(
        location=location, tooltip=tooltip,
        icon=folium.DivIcon(
            html=arrowhead_html(bearing_deg, color, size),
            icon_size=(w, h), icon_anchor=(w // 2, h // 2),
        )
    ).add_to(group)

def add_arrows(coords, color, group, tip=''):
    n = len(coords)
    if n < 2:
        return
    place_arrowhead(coords[0],    stable_bearing(coords, 0.0,  0.15), color, group, f"START {tip}")
    place_arrowhead(coords[n//2], stable_bearing(coords, 0.40, 0.80), color, group, tip)
    place_arrowhead(coords[-1],   stable_bearing(coords, 0.85, 1.0),  color, group, f"END {tip}")

# ── GPS outlier filter ────────────────────────────────────────────────────────

def largest_segment(coords, max_step_deg=0.05):
    """
    Split a coordinate list at GPS jumps larger than max_step_deg and
    return the longest contiguous segment.
    """
    if len(coords) < 2:
        return coords
    segments, current = [], [coords[0]]
    for prev, curr in zip(coords, coords[1:]):
        if abs(curr[0]-prev[0]) > max_step_deg or abs(curr[1]-prev[1]) > max_step_deg:
            segments.append(current)
            current = [curr]
        else:
            current.append(curr)
    segments.append(current)
    return max(segments, key=len)

# ── Shared skip logic ─────────────────────────────────────────────────────────

SKIP_PREFIXES = (
    'TAKE OFF', 'TAKEOFF', 'TEST FIRE', 'LAND', 'KL ', 'ALTITUDE',
    'FRAME CHECK', 'LOW CLOUD', 'SKIPPING', 'SAW GROUP', 'CHECK FRAME',
    'FUEL', 'CLOUDS', 'LAST PASS', 'OBSERVERS', 'PAKT', 'PASI LANDING',
    'ADD ', 'CHECK FOR', 'PREVIOUS', 'ACCIDENTAL', 'COUNTERS', 'ALL ANIMALS',
    'LOOK OUT', 'DISTURBANCE', 'START OF', 'WEST SIDE', 'ANIMALS ON',
    'PASS 1 ', 'PASS 2 ', 'PASS 3 ', 'PASS 4 ', 'PASS 5 ', 'PASS 6 ',
    'PASS 7 ', 'PASS 8 ', 'PASS 9 ', 'PASS 183', 'PASS 2\n',
    'NEW SITE', '1529', 'HIT OUR', '10 JUMPER', '2 JUMPER',
    # ALI-specific operational notes
    'PORT -', 'STAR -', 'BB ', 'REFORMATTED', 'SET APERTURE',
    'DISREGARD', 'NO OPENING', 'SILVER BOX', 'PHOTO PASS OF',
)

def is_operational(comment):
    return comment.upper().strip().startswith(SKIP_PREFIXES)

# ── GOA 2021 site label parsing ───────────────────────────────────────────────

NAME_OVERRIDES_GOA = {'203': 'Ushagat/SW'}

_labels_goa_2021 = {}

def get_site_label_goa_2021(comment):
    c = re.sub(r'^[^A-Za-z0-9]+', '', comment.strip()).strip()
    if re.match(r'^FORRESTER\s+PASS', c, re.IGNORECASE):
        _labels_goa_2021.setdefault('FORRESTER', 'Forrester')
        return _labels_goa_2021['FORRESTER']
    m = re.match(
        r'^(?:SSL?)?(\d+)[A-Z]?\s+(.+?)'
        r'(?:\s+(?:PASS|ABORTED|ONE\s+PASS|NO\s+ANIMALS|OBSV|VISUAL|PHOTO|PASS\s+ONE)'
        r'|\s*[-–]\s*|\s*$)',
        c, re.IGNORECASE
    )
    if not m:
        return None
    sid  = m.group(1)
    name = m.group(2).strip()
    name = re.sub(r'\s+[A-Z]\s+(?:TO|AND)\s+[A-Z]\s*$', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+[A-Z]\s*$', '', name).strip()
    name = re.sub(r'\s+\d+\s*$', '', name).strip()
    if sid in NAME_OVERRIDES_GOA:
        name = NAME_OVERRIDES_GOA[sid]
    else:
        name = name.title()
    if not name:
        return None
    _labels_goa_2021.setdefault(sid, f"{name} ({sid})")
    return _labels_goa_2021[sid]

# ── GOA 2024 site label parsing ───────────────────────────────────────────────

_labels_goa_2024 = {}

def get_site_label_goa_2024(site_id_raw, site_name_raw):
    """Return label like 'Shaw (233)' from raw CSV site_id and site_name."""
    m = re.match(r'(\d+)', str(site_id_raw).strip())
    parent_id = m.group(1) if m else str(site_id_raw).strip()
    if parent_id in _labels_goa_2024:
        return _labels_goa_2024[parent_id]
    if parent_id in NAME_OVERRIDES_GOA:
        name = NAME_OVERRIDES_GOA[parent_id]
    else:
        name = str(site_name_raw).strip()
        name = re.sub(r'\s+[A-Za-z]\s+to\s+[A-Za-z]\s*$', '', name, flags=re.IGNORECASE).strip()
        name = re.sub(r'/[A-Za-z]\s*$', '', name).strip().strip('/')
        name = name.title()
    label = f"{name} ({parent_id})"
    _labels_goa_2024[parent_id] = label
    return label

# ── ALI (2022/2023) site label parsing ────────────────────────────────────────
# ALI flight logs use "SL" prefix (e.g. "SL398 SILAK") or plain numbers
# (e.g. "337 UNALASKA/BISHOP POINT"). The GOA regex uses "SSL?" which matches
# "SS" or "SSL" but not "SL", so ALI needs its own regex with "SL?".

_labels_ali = {}   # shared across 2022 and 2023 for consistent naming

def get_site_label_ali(comment):
    c = re.sub(r'^[^A-Za-z0-9]+', '', comment.strip()).strip()
    m = re.match(
        r'^(?:SL?)?(\d+)[A-Za-z]?\s+(.+?)'
        r'(?:\s+(?:PASS|ABORTED|ONE\s+PASS|NO\s+ANIMALS|\d+\s+ANIMALS?|OBSV|VISUAL|'
        r'PHOTO|PASS\s+ONE|FIRST\s+PHOTO|TOOK\b|COMMENT\b)'
        r'|\s*[-–;]\s*|\s*$)',
        c, re.IGNORECASE
    )
    if not m:
        return None
    sid  = m.group(1)
    name = m.group(2).strip()
    # Strip "TO SL326A " type cross-site prefix
    name = re.sub(r'^TO\s+(?:SL?)?\d*[A-Za-z]?\s*', '', name, flags=re.IGNORECASE).strip()
    # Strip trailing instruction phrases ("ADD FOUR ON RIGHT", "TOOK PASS", etc.)
    name = re.sub(r'\s+(?:ADD|COUNT|GOT\s+THEM|WILL\s+DROP)\b.*', '', name, flags=re.IGNORECASE).strip()
    # Strip harbor-seal cross-reference ("TO HS1910 AND HS1913")
    name = re.sub(r'\s+TO\s+HS\d+.*', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+[A-Z]\s+(?:TO|AND)\s+[A-Z]\s*$', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+[A-Z]\s*$', '', name).strip()
    name = re.sub(r'\s+\d+\s*$', '', name).strip()
    name = re.sub(r'\s+ANIMALS?\s*$', '', name, flags=re.IGNORECASE).strip()
    name = name.title()
    if not name or len(name) < 2:
        return None
    _labels_ali.setdefault(sid, f"{name} ({sid})")
    return _labels_ali[sid]

# ── Generic X/C-row file loader (xlsx and ALI csv) ────────────────────────────

def load_xc_xlsx(filepath, comment_col=30):
    """
    Load an xlsx flight log using X/C row convention.
    X rows supply decimal-degree coords (col 4=lat, col 5=lon).
    C rows supply the pass comment at comment_col.
    Returns list of (date_str, comment, [(lat,lon),...]).
    """
    date_str = Path(filepath).stem[:8]
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    result = []
    current_x = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t, lat, lon = row[0], row[4], row[5]
        comment = row[comment_col] if len(row) > comment_col else None
        if t == 'X' and lat and lon:
            try:
                current_x.append((float(lat), float(lon)))
            except (TypeError, ValueError):
                pass
        elif t == 'C':
            if comment and current_x:
                result.append((date_str, str(comment).strip(), list(current_x)))
            current_x = []
    return result

def load_xc_csv(filepath, comment_col=28):
    """
    Load an ALI csv flight log using X/C row convention.
    X rows supply decimal-degree coords (col 4=lat, col 5=lon).
    C rows supply the pass comment at comment_col.
    Returns list of (date_str, comment, [(lat,lon),...]).
    """
    date_str = Path(filepath).stem[:8]
    result = []
    current_x = []
    with open(filepath, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if not row:
                continue
            t = row[0]
            lat_raw = row[4] if len(row) > 4 else ''
            lon_raw = row[5] if len(row) > 5 else ''
            comment = row[comment_col] if len(row) > comment_col else ''
            if t == 'X' and lat_raw and lon_raw:
                try:
                    current_x.append((float(lat_raw), float(lon_raw)))
                except (TypeError, ValueError):
                    pass
            elif t == 'C':
                if comment and current_x:
                    result.append((date_str, comment.strip(), list(current_x)))
                current_x = []
    return result

# ── Load GOA 2021 passes (xlsx, X/C rows, plain numeric IDs) ──────────────────

passes_by_site_goa_2021 = defaultdict(list)

xlsx_files_goa_2021 = sorted([
    f for f in glob.glob('flightlogs/**/2021/*.xlsx', recursive=True)
    if 'LOGSummary' not in f and 'ASSLAP' not in f
])
for filepath in xlsx_files_goa_2021:
    for date_str, comment, coords in load_xc_xlsx(filepath, comment_col=30):
        if is_operational(comment):
            continue
        label = get_site_label_goa_2021(comment)
        if label:
            passes_by_site_goa_2021[label].append({
                'date': date_str, 'comment': comment,
                'coords': largest_segment(coords),
            })

print(f"GOA 2021: {sum(len(v) for v in passes_by_site_goa_2021.values())} passes across "
      f"{len(passes_by_site_goa_2021)} sites from {len(xlsx_files_goa_2021)} xlsx files.")

# ── Load GOA 2024 passes (NMEA csv, $X rows) ──────────────────────────────────

passes_by_site_goa_2024 = defaultdict(list)

csv_files_goa_2024 = sorted(glob.glob('flightlogs/**/2024/*.csv', recursive=True))
for filepath in csv_files_goa_2024:
    m = re.search(r'(\d{4}-\d{2}-\d{2})', filepath)
    date_str = m.group(1).replace('-', '') if m else 'unknown'
    file_passes = defaultdict(list)
    file_names  = {}
    with open(filepath, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if len(row) < 31 or row[0] != '$X':
                continue
            site_name = row[29].strip()
            site_id   = row[27].strip()
            pass_num  = row[30].strip()
            if not (site_name and site_id):
                continue
            lat_val, lat_ns = row[6], row[7]
            lon_val, lon_ew = row[8], row[9]
            if not (lat_val and lon_val and lat_ns and lon_ew):
                continue
            try:
                lat = nmea_to_dd(lat_val, lat_ns)
                lon = nmea_to_dd(lon_val, lon_ew)
            except (ValueError, ZeroDivisionError):
                continue
            file_passes[(site_id, pass_num)].append((lat, lon))
            file_names.setdefault(site_id, site_name)
    for (site_id, pass_num), coords in file_passes.items():
        coords = largest_segment(coords)
        if not coords:
            continue
        site_name = file_names.get(site_id, site_id)
        label = get_site_label_goa_2024(site_id, site_name)
        passes_by_site_goa_2024[label].append({
            'date': date_str,
            'comment': f"{site_name} pass {pass_num}",
            'coords': coords,
        })

print(f"GOA 2024: {sum(len(v) for v in passes_by_site_goa_2024.values())} passes across "
      f"{len(passes_by_site_goa_2024)} sites from {len(csv_files_goa_2024)} csv files.")

# ── Load ALI passes (xlsx + csv, X/C rows, SL-prefixed IDs) ──────────────────

def load_ali_year(year_str):
    """Load all ALI passes for a given year from xlsx and csv files."""
    passes = defaultdict(list)
    xlsx_files = sorted([
        f for f in glob.glob(f'flightlogs/**/Aleutian Islands/**/{year_str}/*.xlsx', recursive=True)
        if 'ASSLAP' not in f and 'LOGSummary' not in f
    ])
    # Also match without nested region subdirectory
    xlsx_files += sorted([
        f for f in glob.glob(f'flightlogs/**/{year_str}/*.xlsx', recursive=True)
        if 'Aleutian' in f and 'ASSLAP' not in f and 'LOGSummary' not in f
        and f not in xlsx_files
    ])
    for fp in xlsx_files:
        for date_str, comment, coords in load_xc_xlsx(fp, comment_col=30):
            if is_operational(comment):
                continue
            label = get_site_label_ali(comment)
            if label:
                passes[label].append({
                    'date': date_str, 'comment': comment,
                    'coords': largest_segment(coords),
                })

    csv_files = sorted([
        f for f in glob.glob(f'flightlogs/**/{year_str}/*.csv', recursive=True)
        if 'Aleutian' in f
    ])
    for fp in csv_files:
        for date_str, comment, coords in load_xc_csv(fp, comment_col=28):
            if is_operational(comment):
                continue
            label = get_site_label_ali(comment)
            if label:
                passes[label].append({
                    'date': date_str, 'comment': comment,
                    'coords': largest_segment(coords),
                })
    # Normalize antimeridian GPS sign flips — all ALI sites are in western hemisphere
    for pass_list in passes.values():
        all_lons = [c[1] for p in pass_list for c in p['coords']]
        if all_lons and max(all_lons) - min(all_lons) > 180:
            for p in pass_list:
                p['coords'] = [(lat, -lon if lon > 0 else lon) for lat, lon in p['coords']]

    n_xlsx = len(xlsx_files)
    n_csv  = len(csv_files)
    print(f"ALI {year_str}: {sum(len(v) for v in passes.values())} passes across "
          f"{len(passes)} sites from {n_xlsx} xlsx + {n_csv} csv files.")
    return passes

passes_by_site_ali_2022 = load_ali_year('2022')
passes_by_site_ali_2023 = load_ali_year('2023')

# ── Load ASSLAP log notes ──────────────────────────────────────────────────────

def load_log_notes(filepath, col_date=0, col_mml=1, col_pass=6, col_desc=9, sheet_name=None):
    """
    Read an ASSLAP LOGSummary xlsx and return:
      {parent_id_str: [(date_str, pass_num, description), ...]}
    Rows with no pass number or no description are skipped.
    sheet_name: if provided, read that sheet instead of the active sheet.
    """
    notes = defaultdict(list)
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(col_date, col_mml, col_pass, col_desc):
                continue
            date_val = row[col_date]
            mml_id   = row[col_mml]
            pass_num = row[col_pass]
            desc     = row[col_desc]
            if pass_num is None or not desc:
                continue
            m = re.match(r'(\d+)', str(mml_id).strip())
            if not m:
                continue
            parent_id = m.group(1)
            date_str = date_val.strftime('%m/%d') if hasattr(date_val, 'strftime') else str(date_val)
            try:
                pn = int(pass_num)
            except (ValueError, TypeError):
                pn = str(pass_num)
            notes[parent_id].append((date_str, pn, str(desc).strip()))
    except FileNotFoundError:
        print(f"  (log summary not found at {filepath} — skipping)")
    return {k: sorted(v, key=lambda x: (x[0], str(x[1]))) for k, v in notes.items()}

# GOA 2024 ASSLAP (col_date=0, col_mml=1, col_pass=6, col_desc=9)
_asslap_goa_2024 = glob.glob('flightlogs/**/2024/*ASSLAP*.xlsx', recursive=True)
log_notes_goa_2024 = load_log_notes(_asslap_goa_2024[0]) if _asslap_goa_2024 else {}
print(f"GOA 2024 log notes: {len(log_notes_goa_2024)} sites.")

# ALI 2022 ASSLAP (col_date=2, col_mml=3, col_pass=7, col_desc=10; single sheet)
_asslap_ali_2022 = glob.glob('flightlogs/**/2022/*ASSLAP*.xlsx', recursive=True)
log_notes_ali_2022 = (
    load_log_notes(_asslap_ali_2022[0], col_date=2, col_mml=3, col_pass=7, col_desc=10)
    if _asslap_ali_2022 else {}
)
print(f"ALI 2022 log notes: {len(log_notes_ali_2022)} sites.")

# ALI 2023 ASSLAP (survey data is in 'ASSLAP23_SurveySites' sheet, not the active sheet)
_asslap_ali_2023 = glob.glob('flightlogs/**/2023/*ASSLAP*.xlsx', recursive=True)
log_notes_ali_2023 = (
    load_log_notes(_asslap_ali_2023[0], col_date=2, col_mml=3, col_pass=7, col_desc=10,
                   sheet_name='ASSLAP23_SurveySites')
    if _asslap_ali_2023 else {}
)
print(f"ALI 2023 log notes: {len(log_notes_ali_2023)} sites.")

# ── Match site photos ──────────────────────────────────────────────────────────

def find_photo(label, photo_dir):
    safe = re.sub(r'[\\/:*?"<>|]', '_', label)
    for ext in ('png', 'jpg', 'jpeg'):
        p = Path(photo_dir) / f"{safe}.{ext}"
        if p.exists():
            return str(p).replace('\\', '/')
    return None

site_photos_goa_2021 = {l: find_photo(l, 'photos/2021') for l in passes_by_site_goa_2021 if find_photo(l, 'photos/2021')}
site_photos_goa_2024 = {l: find_photo(l, 'photos/2024') for l in passes_by_site_goa_2024 if find_photo(l, 'photos/2024')}
site_photos_ali_2022 = {l: find_photo(l, 'photos/2022') for l in passes_by_site_ali_2022 if find_photo(l, 'photos/2022')}
site_photos_ali_2023 = {l: find_photo(l, 'photos/2023') for l in passes_by_site_ali_2023 if find_photo(l, 'photos/2023')}

print(f"Photos: GOA 2021={len(site_photos_goa_2021)}, GOA 2024={len(site_photos_goa_2024)}, "
      f"ALI 2022={len(site_photos_ali_2022)}, ALI 2023={len(site_photos_ali_2023)}")

# ── Colour palettes ────────────────────────────────────────────────────────────

COLORS = [
    '#e6194b','#3cb44b','#ffe119','#4363d8','#f58231',
    '#911eb4','#42d4f4','#f032e6','#bfef45','#fabed4',
    '#469990','#dcbeff','#9a6324','#800000','#aaffc3',
    '#808000','#ffd8b1','#000075','#a9a9a9','#e6beff',
]
LIGHT_COLORS = {'#ffe119','#bfef45','#42d4f4','#fabed4','#aaffc3','#ffd8b1','#e6beff'}

# Same site across years gets same colour within a region
all_goa_labels = sorted(set(passes_by_site_goa_2021) | set(passes_by_site_goa_2024))
all_ali_labels = sorted(set(passes_by_site_ali_2022) | set(passes_by_site_ali_2023))
color_map_goa = {l: COLORS[i % len(COLORS)] for i, l in enumerate(all_goa_labels)}
color_map_ali = {l: COLORS[i % len(COLORS)] for i, l in enumerate(all_ali_labels)}

# ── Build map ──────────────────────────────────────────────────────────────────

all_coords = [
    c
    for sites in (passes_by_site_goa_2021, passes_by_site_goa_2024,
                  passes_by_site_ali_2022, passes_by_site_ali_2023)
    for site_passes in sites.values()
    for p in site_passes
    for c in p['coords']
]
min_lat = min(c[0] for c in all_coords)
max_lat = max(c[0] for c in all_coords)
min_lon = min(c[1] for c in all_coords)
max_lon = max(c[1] for c in all_coords)
center  = ((min_lat + max_lat) / 2, (min_lon + max_lon) / 2)

m = folium.Map(location=[57, -155], zoom_start=5, tiles=None)

folium.TileLayer(
    tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
    attr='Esri World Imagery', name='Satellite', overlay=False, control=True,
).add_to(m)
folium.TileLayer(
    tiles='https://server.arcgisonline.com/ArcGIS/rest/services/Reference/World_Boundaries_and_Places/MapServer/tile/{z}/{y}/{x}',
    attr='Esri Labels', name='Labels', overlay=True, control=True, opacity=0.7,
).add_to(m)

DETAIL_ZOOM = 9

def add_site_layers(passes_by_site, site_photos, layer_prefix, color_map, log_notes=None, show=True):
    """
    Add one FeatureGroup per site.
    layer_prefix: e.g. 'GOA 2021', 'ALI 2022' — used as the first token in the
                  layer name so the JS toggle can filter by region+year.
    """
    for site in sorted(passes_by_site):
        color      = color_map[site]
        text_color = '#000' if color in LIGHT_COLORS else '#fff'
        group_name = f"{layer_prefix} | {site}"
        group      = folium.FeatureGroup(name=group_name, show=show)
        site_passes = passes_by_site[site]

        all_site_coords = [c for p in site_passes for c in p['coords']]
        centroid = (
            sum(c[0] for c in all_site_coords) / len(all_site_coords),
            sum(c[1] for c in all_site_coords) / len(all_site_coords),
        )

        # Overview dot — visible at low zoom
        folium.Marker(
            location=centroid,
            tooltip=site,
            icon=folium.DivIcon(
                html=(f'<div data-mtype="overview" '
                      f'style="width:10px;height:10px;border-radius:50%;'
                      f'background:{color};'
                      f'border:2px solid rgba(255,255,255,0.85);'
                      f'box-shadow:0 0 4px rgba(0,0,0,0.7);"></div>'),
                icon_size=(10, 10), icon_anchor=(5, 5),
            ),
        ).add_to(group)

        for pass_num, p in enumerate(site_passes, 1):
            coords = p['coords']
            badge  = f"P{pass_num}"
            tip    = f"{badge} ({p['date']}) — {p['comment']}"

            folium.PolyLine(
                locations=coords, color=color, weight=3, opacity=0.9, tooltip=tip,
            ).add_to(group)

            add_arrows(coords, color, group, tip)

            q = max(0, len(coords) // 4)
            folium.Marker(
                location=coords[q], tooltip=tip,
                icon=folium.DivIcon(
                    html=(f'<div data-mtype="detail" '
                          f'style="background:{color};color:{text_color};'
                          f'border-radius:50%;width:20px;height:20px;line-height:20px;'
                          f'text-align:center;font-size:10px;font-weight:bold;'
                          f'font-family:sans-serif;border:1.5px solid rgba(0,0,0,0.4);'
                          f'box-shadow:1px 1px 3px rgba(0,0,0,0.6);">{badge}</div>'),
                    icon_size=(20, 20), icon_anchor=(10, 10),
                )
            ).add_to(group)

        if site in site_photos:
            site_name_short = re.sub(r'\s*\(\d+\)\s*$', '', site).strip()

            notes_html = ''
            if log_notes:
                pid = re.search(r'\((\d+)\)', site)
                pid = pid.group(1) if pid else None
                entries = log_notes.get(pid, []) if pid else []
                if entries:
                    dates = {d for d, _, _ in entries}
                    show_date = len(dates) > 1
                    rows = ''.join(
                        f'<div style="margin:2px 0;">'
                        f'{"<b>" + d + "</b> — " if show_date else ""}'
                        f'Pass {pn}: {desc}</div>'
                        for d, pn, desc in entries
                    )
                    notes_html = (
                        f'<div style="text-align:left;margin-top:8px;padding-top:6px;'
                        f'border-top:1px solid #ddd;font-size:11px;'
                        f'font-family:sans-serif;color:#333;">'
                        f'{rows}</div>'
                    )

            popup_html = (
                f'<div style="text-align:center;font-family:sans-serif;padding:4px;">'
                f'<b style="font-size:13px;">{site} ({layer_prefix})</b><br>'
                f'<img src="{site_photos[site]}" '
                f'style="max-width:300px;max-height:220px;margin-top:6px;border-radius:4px;">'
                f'{notes_html}'
                f'</div>'
            )
            folium.Marker(
                location=centroid,
                tooltip=f"{site} ({layer_prefix}) — click for photo",
                popup=folium.Popup(popup_html, max_width=320),
                icon=folium.DivIcon(
                    html=(f'<div data-mtype="detail" '
                          f'style="background:{color};color:{text_color};'
                          f'border-radius:4px;padding:2px 6px;'
                          f'font-size:10px;font-weight:bold;font-family:sans-serif;'
                          f'border:1.5px solid rgba(0,0,0,0.4);'
                          f'box-shadow:1px 1px 3px rgba(0,0,0,0.5);white-space:nowrap;">'
                          f'&#128247; {site_name_short}</div>'),
                    icon_size=(len(site_name_short) * 7 + 30, 20),
                    icon_anchor=(-8, 10),
                )
            ).add_to(group)

        group.add_to(m)

# GOA layers shown by default; ALI layers hidden (JS toggle reveals them)
add_site_layers(passes_by_site_goa_2021, site_photos_goa_2021, 'GOA 2021', color_map_goa, show=True)
add_site_layers(passes_by_site_goa_2024, site_photos_goa_2024, 'GOA 2024', color_map_goa,
                log_notes=log_notes_goa_2024, show=True)
add_site_layers(passes_by_site_ali_2022, site_photos_ali_2022, 'ALEU 2022', color_map_ali,
                log_notes=log_notes_ali_2022, show=False)
add_site_layers(passes_by_site_ali_2023, site_photos_ali_2023, 'ALEU 2023', color_map_ali,
                log_notes=log_notes_ali_2023, show=False)

folium.LayerControl(collapsed=True).add_to(m)

# ── Region + year toggle UI ────────────────────────────────────────────────────

toggle_html = """
<style>
  #survey-toggle {
    position: fixed; top: 10px; left: 50%; transform: translateX(-50%);
    z-index: 9999; display: flex; flex-direction: column; align-items: center;
    background: rgba(255,255,255,0.95); border-radius: 8px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.35); overflow: hidden;
  }
  #region-toggle, #year-toggle-row {
    display: flex; width: 100%;
  }
  #survey-toggle button {
    padding: 7px 18px; border: none; cursor: pointer;
    font-family: sans-serif; font-size: 13px; font-weight: bold;
    background: #eee; color: #555; transition: background .15s, color .15s;
    flex: 1;
  }
  #region-toggle button { font-size: 12px; border-bottom: 1px solid #ddd; }
  #region-toggle button.active { background: #1a4a7a; color: #fff; }
  #year-toggle-row button.active { background: #4363d8; color: #fff; }
  #year-toggle-row { border-top: 1px solid #ddd; }

  /* Zoom-based marker visibility via CSS */
  .leaflet-marker-icon:has([data-mtype="detail"])  { display: none; }
  .leaflet-marker-icon:has([data-mtype="overview"]) { display: block; }
  .zoom-detail .leaflet-marker-icon:has([data-mtype="detail"])  { display: block; }
  .zoom-detail .leaflet-marker-icon:has([data-mtype="overview"]) { display: none; }

  #layer-search {
    display: block; width: calc(100% - 16px); margin: 6px 8px 4px;
    padding: 5px 8px; border: 1px solid #ccc; border-radius: 4px;
    font-size: 12px; font-family: sans-serif; box-sizing: border-box;
  }
  #layer-search:focus { outline: none; border-color: #4363d8; }
  .layer-hidden { display: none !important; }
</style>

<div id="survey-toggle">
  <div id="region-toggle">
    <button data-region="goa" class="active" onclick="setRegion('goa')">Gulf of Alaska</button>
    <button data-region="aleu" onclick="setRegion('aleu')">Aleutian Islands</button>
  </div>
  <div id="year-toggle-row">
    <div id="goa-years" style="display:flex;flex:1;">
      <button data-yr="2021" onclick="setYear('2021')">2021</button>
      <button data-yr="both" class="active" onclick="setYear('both')">Both</button>
      <button data-yr="2024" onclick="setYear('2024')">2024</button>
    </div>
    <div id="aleu-years" style="display:none;flex:1;">
      <button data-yr="2022" onclick="setYear('2022')">2022</button>
      <button data-yr="both" class="active" onclick="setYear('both')">Both</button>
      <button data-yr="2023" onclick="setYear('2023')">2023</button>
    </div>
  </div>
</div>

<script>
// ── Region + year toggle ──────────────────────────────────────────────────────
var currentRegion = 'goa';
var currentYear   = {goa: 'both', aleu: 'both'};

function setRegion(r) {
  currentRegion = r;
  document.querySelectorAll('#region-toggle button').forEach(function(b) {
    b.classList.toggle('active', b.getAttribute('data-region') === r);
  });
  document.getElementById('goa-years').style.display  = (r === 'goa')  ? 'flex' : 'none';
  document.getElementById('aleu-years').style.display = (r === 'aleu') ? 'flex' : 'none';
  var yr = currentYear[r];
  document.querySelectorAll('#' + r + '-years button').forEach(function(b) {
    b.classList.toggle('active', b.getAttribute('data-yr') === yr);
  });
  updateLayers();
}

function setYear(yr) {
  currentYear[currentRegion] = yr;
  document.querySelectorAll('#' + currentRegion + '-years button').forEach(function(b) {
    b.classList.toggle('active', b.getAttribute('data-yr') === yr);
  });
  updateLayers();
}

function updateLayers() {
  var r  = currentRegion;
  var yr = currentYear[r];
  document.querySelectorAll('.leaflet-control-layers-overlays label').forEach(function(lbl) {
    var txt = lbl.innerText.trim();
    var inp = lbl.querySelector('input[type="checkbox"]');
    if (!inp) return;
    var isGOA21  = txt.startsWith('GOA 2021 |');
    var isGOA24  = txt.startsWith('GOA 2024 |');
    var isALEU22 = txt.startsWith('ALEU 2022 |');
    var isALEU23 = txt.startsWith('ALEU 2023 |');
    if (!isGOA21 && !isGOA24 && !isALEU22 && !isALEU23) return;
    var inRegion = (r === 'goa') ? (isGOA21 || isGOA24) : (isALEU22 || isALEU23);
    var show;
    if (!inRegion) {
      show = false;
    } else if (yr === 'both') {
      show = true;
    } else if (r === 'goa') {
      show = (yr === '2021' && isGOA21) || (yr === '2024' && isGOA24);
    } else {
      show = (yr === '2022' && isALEU22) || (yr === '2023' && isALEU23);
    }
    if (inp.checked !== show) inp.click();
  });
}

// ── Zoom-based marker visibility ──────────────────────────────────────────────
var DETAIL_ZOOM = 9;
window.addEventListener('load', function() {
  var _map = null;
  for (var k in window) {
    try {
      if (window[k] && typeof window[k].getZoom === 'function' &&
          typeof window[k].on === 'function') { _map = window[k]; break; }
    } catch(e) {}
  }
  if (!_map) return;
  function updateZoomVis() {
    _map.getContainer().classList.toggle('zoom-detail', _map.getZoom() >= DETAIL_ZOOM);
  }
  _map.on('zoom', updateZoomVis);
  _map.on('layeradd', updateZoomVis);
  updateZoomVis();
});

// ── Search box ────────────────────────────────────────────────────────────────
function injectSearch() {
  var overlays = document.querySelector('.leaflet-control-layers-overlays');
  if (!overlays || document.getElementById('layer-search')) return;
  var input = document.createElement('input');
  input.id = 'layer-search';
  input.type = 'text';
  input.placeholder = 'Search sites\u2026';
  overlays.parentNode.insertBefore(input, overlays);
  input.addEventListener('input', function() {
    var q = this.value.trim().toLowerCase();
    document.querySelectorAll('.leaflet-control-layers-overlays label').forEach(function(lbl) {
      var name = lbl.innerText.trim().toLowerCase();
      var stripped = name.replace(/^(goa|aleu) \\d{4} \\| /, '');
      var match = !q || stripped.includes(q) || name.includes(q);
      lbl.classList.toggle('layer-hidden', !match);
    });
  });
}
var _searchObserver = new MutationObserver(function() {
  if (document.querySelector('.leaflet-control-layers-overlays')) injectSearch();
});
window.addEventListener('load', function() {
  _searchObserver.observe(document.body, { childList: true, subtree: true });
  injectSearch();
});
</script>
"""

m.get_root().html.add_child(folium.Element(toggle_html))

# ── Write output ───────────────────────────────────────────────────────────────

outfile = 'index.html'
m.save(outfile)
print(f"\nSaved: {outfile}")

# Summary table
print(f"\n{'Site':<45} {'G21':>4} {'G24':>4} {'A22':>4} {'A23':>4}")
print('-' * 65)
all_labels = sorted(
    set(passes_by_site_goa_2021) | set(passes_by_site_goa_2024) |
    set(passes_by_site_ali_2022) | set(passes_by_site_ali_2023)
)
for site in all_labels:
    g21 = len(passes_by_site_goa_2021.get(site, []))
    g24 = len(passes_by_site_goa_2024.get(site, []))
    a22 = len(passes_by_site_ali_2022.get(site, []))
    a23 = len(passes_by_site_ali_2023.get(site, []))
    print(f"{site:<45} {g21 or '':>4} {g24 or '':>4} {a22 or '':>4} {a23 or '':>4}")
