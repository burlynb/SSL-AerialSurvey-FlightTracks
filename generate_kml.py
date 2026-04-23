"""
generate_kml.py
---------------
Generates KML flight track files for ForeFlight import.

  2021_flighttracks.kml  — GOA 2021
  2024_flighttracks.kml  — GOA 2024
  2022_flighttracks.kml  — ALI 2022
  2023_flighttracks.kml  — ALI 2023

KML design for ForeFlight:
  - One labeled Point per site at track centroid (site number shown on map;
    tap to see site name, year, and per-pass survey notes)
  - One LineString per pass, thick (width 5) with StyleMap highlight
  - Green start dot + bearing arrow at end of each pass
  - ASSLAP pass notes included in popup description where available
"""

import csv
import html as html_mod
import openpyxl
import glob
import os
import re
import shutil
import tempfile
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET
from xml.dom import minidom

# ── Color palette (matches web map) ──────────────────────────────────────────

COLORS_HTML = [
    '#e6194b', '#ff6b6b', '#ffe119', '#4363d8', '#f58231',
    '#911eb4', '#42d4f4', '#f032e6', '#bfef45', '#fabed4',
    '#469990', '#dcbeff', '#9a6324', '#800000', '#c9c9ff',
    '#808000', '#ffd8b1', '#000075', '#a9a9a9', '#e6beff',
]

def to_kml_color(html, alpha='ff'):
    h = html.lstrip('#')
    return f"{alpha}{h[4:6]}{h[2:4]}{h[0:2]}"

# ── Shared skip logic ─────────────────────────────────────────────────────────

SKIP_PREFIXES = (
    'TAKE OFF', 'TAKEOFF', 'TEST FIRE', 'LAND', 'KL ', 'ALTITUDE',
    'FRAME CHECK', 'LOW CLOUD', 'SKIPPING', 'SAW GROUP', 'CHECK FRAME',
    'FUEL', 'CLOUDS', 'LAST PASS', 'OBSERVERS', 'PAKT', 'PASI LANDING',
    'ADD ', 'CHECK FOR', 'PREVIOUS', 'ACCIDENTAL', 'COUNTERS', 'ALL ANIMALS',
    'LOOK OUT', 'DISTURBANCE', 'START OF', 'WEST SIDE', 'ANIMALS ON',
    'PASS 1 ', 'PASS 2 ', 'PASS 3 ', 'PASS 4 ', 'PASS 5 ', 'PASS 6 ',
    'PASS 7 ', 'PASS 8 ', 'PASS 9 ', 'PASS 183', 'NEW SITE',
    '1529', 'HIT OUR', '10 JUMPER', '2 JUMPER',
    'PORT -', 'STAR -', 'BB ', 'REFORMATTED', 'SET APERTURE',
    'DISREGARD', 'NO OPENING', 'SILVER BOX', 'PHOTO PASS OF',
)

def is_operational(comment):
    return comment.upper().strip().startswith(SKIP_PREFIXES)

# ── Site label parsing ────────────────────────────────────────────────────────

NAME_OVERRIDES_GOA = {'203': 'Ushagat/SW'}

_labels_goa_2021 = {}

def get_site_label_goa_2021(comment):
    c = re.sub(r'^[^A-Za-z0-9]+', '', comment.strip()).strip()
    if re.match(r'^FORRESTER\s+PASS', c, re.IGNORECASE):
        _labels_goa_2021.setdefault('FORRESTER', 'Forrester')
        return _labels_goa_2021['FORRESTER']
    m = re.match(
        r'^(?:SSL?)?(\d+)[A-Z]?\s+(.+?)'
        r'(?:\s+(?:PASS|ABORTED|ONE\s+PASS|NO\s+ANIMALS|OBSV|VISUAL|PHOTO|PASS\s+ONE)'
        r'|\s*[-–]\s*|\s*$)',
        c, re.IGNORECASE
    )
    if not m:
        return None
    sid  = m.group(1)
    name = m.group(2).strip()
    name = re.sub(r'\s+[A-Z]\s+(?:TO|AND)\s+[A-Z]\s*$', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+[A-Z]\s*$', '', name).strip()
    name = re.sub(r'\s+\d+\s*$', '', name).strip()
    if sid in NAME_OVERRIDES_GOA:
        name = NAME_OVERRIDES_GOA[sid]
    else:
        name = name.title()
    if not name:
        return None
    _labels_goa_2021.setdefault(sid, f"{name} ({sid})")
    return _labels_goa_2021[sid]

def nmea_to_dd(val, hemisphere):
    v = float(val)
    deg = int(v / 100)
    dd = deg + (v - deg * 100) / 60.0
    if hemisphere.upper() in ('S', 'W'):
        dd = -dd
    return dd

_labels_goa_2024 = {}

def get_site_label_goa_2024(site_id_raw, site_name_raw):
    m = re.match(r'(\d+)', str(site_id_raw).strip())
    parent_id = m.group(1) if m else str(site_id_raw).strip()
    if parent_id in _labels_goa_2024:
        return _labels_goa_2024[parent_id]
    if parent_id in NAME_OVERRIDES_GOA:
        name = NAME_OVERRIDES_GOA[parent_id]
    else:
        name = str(site_name_raw).strip()
        name = re.sub(r'\s+[A-Za-z]\s+to\s+[A-Za-z]\s*$', '', name, flags=re.IGNORECASE).strip()
        name = re.sub(r'/[A-Za-z]\s*$', '', name).strip().strip('/')
        name = name.title()
    label = f"{name} ({parent_id})"
    _labels_goa_2024[parent_id] = label
    return label

_labels_ali = {}

def get_site_label_ali(comment):
    c = re.sub(r'^[^A-Za-z0-9]+', '', comment.strip()).strip()
    m = re.match(
        r'^(?:SL?)?(\d+)[A-Za-z]?\s+(.+?)'
        r'(?:\s+(?:PASS|ABORTED|ONE\s+PASS|NO\s+ANIMALS|\d+\s+ANIMALS?|OBSV|VISUAL|'
        r'PHOTO|PASS\s+ONE|FIRST\s+PHOTO|TOOK\b|COMMENT\b)'
        r'|\s*[-–;]\s*|\s*$)',
        c, re.IGNORECASE
    )
    if not m:
        return None
    sid  = m.group(1)
    name = m.group(2).strip()
    name = re.sub(r'^TO\s+(?:SL?)?\d*[A-Za-z]?\s*', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+(?:ADD|COUNT|GOT\s+THEM|WILL\s+DROP)\b.*', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+TO\s+HS\d+.*', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+[A-Z]\s+(?:TO|AND)\s+[A-Z]\s*$', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'\s+[A-Z]\s*$', '', name).strip()
    name = re.sub(r'\s+\d+\s*$', '', name).strip()
    name = re.sub(r'\s+ANIMALS?\s*$', '', name, flags=re.IGNORECASE).strip()
    name = name.title()
    if not name or len(name) < 2:
        return None
    _labels_ali.setdefault(sid, f"{name} ({sid})")
    return _labels_ali[sid]

# ── Generic X/C-row loaders ───────────────────────────────────────────────────

def load_xc_xlsx(filepath, comment_col=30):
    date_str = Path(filepath).stem[:8]
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    result, current_x = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t, lat, lon = row[0], row[4], row[5]
        comment = row[comment_col] if len(row) > comment_col else None
        if t == 'X' and lat and lon:
            try:
                current_x.append((float(lat), float(lon)))
            except (TypeError, ValueError):
                pass
        elif t == 'C':
            if comment and current_x:
                result.append((date_str, str(comment).strip(), list(current_x)))
            current_x = []
    return result

def load_xc_csv_ali(filepath, comment_col=28):
    date_str = Path(filepath).stem[:8]
    result, current_x = [], []
    with open(filepath, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if not row:
                continue
            t = row[0]
            lat_raw = row[4] if len(row) > 4 else ''
            lon_raw = row[5] if len(row) > 5 else ''
            comment = row[comment_col] if len(row) > comment_col else ''
            if t == 'X' and lat_raw and lon_raw:
                try:
                    current_x.append((float(lat_raw), float(lon_raw)))
                except (TypeError, ValueError):
                    pass
            elif t == 'C':
                if comment and current_x:
                    result.append((date_str, comment.strip(), list(current_x)))
                current_x = []
    return result

# ── GPS outlier filter ────────────────────────────────────────────────────────

def largest_segment(coords, max_step_deg=0.05):
    if len(coords) < 2:
        return coords
    segments, current = [], [coords[0]]
    for prev, curr in zip(coords, coords[1:]):
        if abs(curr[0]-prev[0]) > max_step_deg or abs(curr[1]-prev[1]) > max_step_deg:
            segments.append(current)
            current = [curr]
        else:
            current.append(curr)
    segments.append(current)
    return max(segments, key=len)

# ── Antimeridian normalization ────────────────────────────────────────────────

def fix_antimeridian(passes):
    for pass_list in passes.values():
        all_lons = [c[1] for p in pass_list for c in p['coords']]
        if all_lons and max(all_lons) - min(all_lons) > 180:
            for p in pass_list:
                p['coords'] = [(lat, -lon if lon > 0 else lon) for lat, lon in p['coords']]
    return passes

# ── Load passes ───────────────────────────────────────────────────────────────

passes_by_site_goa_2021 = defaultdict(list)
for fp in sorted([f for f in glob.glob('flightlogs/**/2021/*.xlsx', recursive=True)
                  if 'LOGSummary' not in f and 'ASSLAP' not in f]):
    for date_str, comment, coords in load_xc_xlsx(fp, comment_col=30):
        if is_operational(comment):
            continue
        label = get_site_label_goa_2021(comment)
        if label:
            passes_by_site_goa_2021[label].append(
                {'date': date_str, 'comment': comment, 'coords': largest_segment(coords)})
print(f"GOA 2021: {sum(len(v) for v in passes_by_site_goa_2021.values())} passes / "
      f"{len(passes_by_site_goa_2021)} sites.")

passes_by_site_goa_2024 = defaultdict(list)
for fp in sorted(glob.glob('flightlogs/**/2024/*.csv', recursive=True)):
    m = re.search(r'(\d{4}-\d{2}-\d{2})', fp)
    date_str = m.group(1).replace('-', '') if m else 'unknown'
    file_passes = defaultdict(list)
    file_names  = {}
    with open(fp, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if len(row) < 31 or row[0] != '$X':
                continue
            site_name = row[29].strip(); site_id = row[27].strip(); pass_num = row[30].strip()
            if not (site_name and site_id):
                continue
            try:
                lat = nmea_to_dd(row[6], row[7]); lon = nmea_to_dd(row[8], row[9])
            except (ValueError, ZeroDivisionError):
                continue
            file_passes[(site_id, pass_num)].append((lat, lon))
            file_names.setdefault(site_id, site_name)
    for (site_id, pass_num), coords in file_passes.items():
        coords = largest_segment(coords)
        if not coords:
            continue
        label = get_site_label_goa_2024(site_id, file_names.get(site_id, site_id))
        passes_by_site_goa_2024[label].append({
            'date': date_str,
            'comment': f"{file_names.get(site_id, site_id)} pass {pass_num}",
            'coords': coords,
        })
print(f"GOA 2024: {sum(len(v) for v in passes_by_site_goa_2024.values())} passes / "
      f"{len(passes_by_site_goa_2024)} sites.")

def load_ali_year(year_str):
    passes = defaultdict(list)
    for fp in sorted([f for f in glob.glob(f'flightlogs/**/{year_str}/*.xlsx', recursive=True)
                      if 'Aleutian' in f and 'ASSLAP' not in f and 'LOGSummary' not in f]):
        for date_str, comment, coords in load_xc_xlsx(fp, comment_col=30):
            if is_operational(comment):
                continue
            label = get_site_label_ali(comment)
            if label:
                passes[label].append(
                    {'date': date_str, 'comment': comment, 'coords': largest_segment(coords)})
    for fp in sorted([f for f in glob.glob(f'flightlogs/**/{year_str}/*.csv', recursive=True)
                      if 'Aleutian' in f]):
        for date_str, comment, coords in load_xc_csv_ali(fp, comment_col=28):
            if is_operational(comment):
                continue
            label = get_site_label_ali(comment)
            if label:
                passes[label].append(
                    {'date': date_str, 'comment': comment, 'coords': largest_segment(coords)})
    fix_antimeridian(passes)
    print(f"ALI {year_str}: {sum(len(v) for v in passes.values())} passes / {len(passes)} sites.")
    return passes

passes_by_site_ali_2022 = load_ali_year('2022')
passes_by_site_ali_2023 = load_ali_year('2023')

# ── ASSLAP log notes ──────────────────────────────────────────────────────────

def load_log_notes(filepath, col_date=0, col_mml=1, col_pass=6, col_desc=9, sheet_name=None):
    notes = defaultdict(list)
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(col_date, col_mml, col_pass, col_desc):
                continue
            mml_id = row[col_mml]; pass_num = row[col_pass]; desc = row[col_desc]
            if pass_num is None or not desc:
                continue
            m = re.match(r'(\d+)', str(mml_id).strip()) if mml_id else None
            if not m:
                continue
            pid = m.group(1)
            date_val = row[col_date]
            date_str = date_val.strftime('%m/%d') if hasattr(date_val, 'strftime') else str(date_val)
            try:
                pn = int(pass_num)
            except (ValueError, TypeError):
                pn = str(pass_num)
            notes[pid].append((date_str, pn, str(desc).strip()))
    except FileNotFoundError:
        print(f"  (ASSLAP not found: {filepath})")
    return {k: sorted(v, key=lambda x: (x[0], str(x[1]))) for k, v in notes.items()}

def load_log_notes_by_name(filepath, col_date=0, col_name=3, col_pass=6, col_desc=10):
    """Load 2021-style ASSLAP where rows have site names (not MML IDs). Returns dict keyed
    by normalized SITENAME (uppercase, alphanumeric only)."""
    notes = defaultdict(list)
    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        shutil.copy2(filepath, tmp_path)
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(col_date, col_name, col_pass, col_desc):
                continue
            name_val = row[col_name]; pass_num = row[col_pass]; desc = row[col_desc]
            if not name_val or pass_num is None or not desc:
                continue
            key = re.sub(r'[^A-Z0-9]', '', str(name_val).upper())
            date_val = row[col_date]
            date_str = date_val.strftime('%m/%d') if hasattr(date_val, 'strftime') else str(date_val)
            try:
                pn = int(pass_num)
            except (ValueError, TypeError):
                pn = str(pass_num)
            notes[key].append((date_str, pn, str(desc).strip()))
    except FileNotFoundError:
        print(f"  (ASSLAP not found: {filepath})")
    except Exception as e:
        print(f"  (ASSLAP error: {e})")
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
    return {k: sorted(v, key=lambda x: (x[0], str(x[1]))) for k, v in notes.items()}

def match_notes_to_ids(notes_by_name, passes_by_site):
    """Re-key a name-keyed notes dict by numeric site ID by matching against site labels.
    Falls back to prefix match (e.g. 'CAPESTELIASB' vs 'CAPESTELIAS') when both keys
    are at least 8 chars, preventing short-key false matches like MIDDLE vs MIDDLETON."""
    MIN_PREFIX = 8
    result = {}
    for site in passes_by_site:
        m = re.search(r'\((\d+)\)', site)
        if not m:
            continue
        sid = m.group(1)
        name_part = re.sub(r'\s*\(\d+\)\s*$', '', site).strip()
        key = re.sub(r'[^A-Z0-9]', '', name_part.upper())
        if key in notes_by_name:
            result[sid] = notes_by_name[key]
            continue
        merged = []
        for n_key, n_notes in notes_by_name.items():
            if len(key) >= MIN_PREFIX and len(n_key) >= MIN_PREFIX:
                if key.startswith(n_key) or n_key.startswith(key):
                    merged.extend(n_notes)
        if merged:
            result[sid] = sorted(merged, key=lambda x: (x[0], str(x[1])))
    return result

_asslap_goa_2024 = glob.glob('flightlogs/**/2024/*ASSLAP*.xlsx', recursive=True)
log_notes_goa_2024 = load_log_notes(_asslap_goa_2024[0]) if _asslap_goa_2024 else {}
print(f"GOA 2024 log notes: {len(log_notes_goa_2024)} sites.")

_asslap_ali_2022 = glob.glob('flightlogs/**/2022/*ASSLAP*.xlsx', recursive=True)
log_notes_ali_2022 = (
    load_log_notes(_asslap_ali_2022[0], col_date=2, col_mml=3, col_pass=7, col_desc=10)
    if _asslap_ali_2022 else {}
)
print(f"ALI 2022 log notes: {len(log_notes_ali_2022)} sites.")

_asslap_ali_2023 = glob.glob('flightlogs/**/2023/*ASSLAP*.xlsx', recursive=True)
log_notes_ali_2023 = (
    load_log_notes(_asslap_ali_2023[0], col_date=2, col_mml=3, col_pass=7, col_desc=10,
                   sheet_name='ASSLAP23_SurveySites')
    if _asslap_ali_2023 else {}
)
print(f"ALI 2023 log notes: {len(log_notes_ali_2023)} sites.")

_asslap_goa_2021 = glob.glob('flightlogs/**/2021/*ASSLAP*.xlsx', recursive=True)
if _asslap_goa_2021:
    _notes_by_name_2021 = load_log_notes_by_name(_asslap_goa_2021[0])
    log_notes_goa_2021 = match_notes_to_ids(_notes_by_name_2021, passes_by_site_goa_2021)
else:
    log_notes_goa_2021 = {}
print(f"GOA 2021 log notes: {len(log_notes_goa_2021)} sites.")

# ── Utilities ─────────────────────────────────────────────────────────────────

def fmt_date(d):
    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    s = str(d)
    if len(s) == 8 and s.isdigit():
        return f"{months[int(s[4:6])-1]} {int(s[6:8])}"
    return s

def make_description(site, year, region, site_passes, log_notes):
    pid = re.search(r'\((\d+)\)', site)
    pid = pid.group(1) if pid else None
    notes_list = log_notes.get(pid, []) if log_notes and pid else []
    notes_by_pass = {}
    for _, pn, desc in notes_list:
        try:
            notes_by_pass[int(pn)] = desc
        except (ValueError, TypeError):
            pass

    n = len(site_passes)
    lines = [
        f'<b>{site}</b>',
        f'{year} — {region}',
        f'<b>{n} pass{"es" if n != 1 else ""}</b>',
        '',
    ]
    for i, p in enumerate(site_passes, 1):
        date_label = fmt_date(p['date'])
        note = notes_by_pass.get(i)
        if note:
            lines.append(f'<b>Pass {i}</b> ({date_label}): {note}')
        else:
            lines.append(f'<b>Pass {i}</b> ({date_label})')
    return '<br/>'.join(lines)

# ── KML builder ───────────────────────────────────────────────────────────────

def build_kml(passes_by_site, year, region, log_notes=None):
    kml = ET.Element('kml', xmlns='http://www.opengis.net/kml/2.2')
    doc = ET.SubElement(kml, 'Document')
    ET.SubElement(doc, 'name').text = f"{year} SSL Aerial Survey — {region}"
    ET.SubElement(doc, 'description').text = (
        f"Steller sea lion aerial survey flight tracks, {region} {year}. "
        f"Tap a site label to see site name and pass notes."
    )

    # ── Shared start-dot style (green) ────────────────────────────────────────
    st = ET.SubElement(doc, 'Style', id='startDot')
    ic = ET.SubElement(st, 'IconStyle')
    ET.SubElement(ic, 'color').text = 'ff00bb00'
    ET.SubElement(ic, 'scale').text = '0.65'
    ico = ET.SubElement(ic, 'Icon')
    ET.SubElement(ico, 'href').text = 'http://maps.google.com/mapfiles/kml/shapes/shaded_dot.png'
    ET.SubElement(ET.SubElement(st, 'LabelStyle'), 'scale').text = '0'

    # ── Shared number-label style (text only, no icon) ────────────────────────
    st_num = ET.SubElement(doc, 'Style', id='numLabel')
    ET.SubElement(ET.SubElement(st_num, 'IconStyle'), 'scale').text = '0'
    num_lb = ET.SubElement(st_num, 'LabelStyle')
    ET.SubElement(num_lb, 'color').text = 'ffffffff'
    ET.SubElement(num_lb, 'scale').text = '0.8'

    # ── Per-color styles: line normal/highlight + site-label normal/highlight + StyleMaps ──
    for i, html_color in enumerate(COLORS_HTML):
        kc = to_kml_color(html_color)

        def line_style(sid, width):
            s = ET.SubElement(doc, 'Style', id=sid)
            ls = ET.SubElement(s, 'LineStyle')
            ET.SubElement(ls, 'color').text = kc
            ET.SubElement(ls, 'width').text = str(width)
            ET.SubElement(ET.SubElement(s, 'LabelStyle'), 'scale').text = '0'

        def dot_style(sid, dot_scale, label_scale):
            s = ET.SubElement(doc, 'Style', id=sid)
            ics = ET.SubElement(s, 'IconStyle')
            ET.SubElement(ics, 'color').text = kc
            ET.SubElement(ics, 'scale').text = str(dot_scale)
            ico2 = ET.SubElement(ics, 'Icon')
            ET.SubElement(ico2, 'href').text = 'http://maps.google.com/mapfiles/kml/shapes/shaded_dot.png'
            lbs = ET.SubElement(s, 'LabelStyle')
            ET.SubElement(lbs, 'color').text = 'ffffffff'
            ET.SubElement(lbs, 'scale').text = str(label_scale)

        def stylemap(smid, normal_id, highlight_id):
            sm = ET.SubElement(doc, 'StyleMap', id=smid)
            pn = ET.SubElement(sm, 'Pair')
            ET.SubElement(pn, 'key').text = 'normal'
            ET.SubElement(pn, 'styleUrl').text = f'#{normal_id}'
            ph = ET.SubElement(sm, 'Pair')
            ET.SubElement(ph, 'key').text = 'highlight'
            ET.SubElement(ph, 'styleUrl').text = f'#{highlight_id}'

        line_style(f'ln{i}', 5)
        line_style(f'lh{i}', 8)
        dot_style(f'sn{i}', 1.0, 0)
        dot_style(f'sh{i}', 1.4, 0)
        stylemap(f'line{i}', f'ln{i}', f'lh{i}')
        stylemap(f'site{i}', f'sn{i}', f'sh{i}')

    # ── Site folders ──────────────────────────────────────────────────────────
    for idx, site in enumerate(sorted(passes_by_site)):
        ci = idx % len(COLORS_HTML)
        kc = to_kml_color(COLORS_HTML[ci])
        site_passes = passes_by_site[site]

        m = re.search(r'\((\d+)\)', site)
        site_num = m.group(1) if m else site

        all_coords = [c for p in site_passes for c in p['coords']]
        cen_lat = sum(c[0] for c in all_coords) / len(all_coords)
        cen_lon = sum(c[1] for c in all_coords) / len(all_coords)

        desc = make_description(site, year, region, site_passes, log_notes)

        folder = ET.SubElement(doc, 'Folder')
        ET.SubElement(folder, 'name').text = site

        # Number text label — "246" visible on map, no popup
        pm_num = ET.SubElement(folder, 'Placemark')
        ET.SubElement(pm_num, 'name').text = site_num
        ET.SubElement(pm_num, 'styleUrl').text = '#numLabel'
        pt_num = ET.SubElement(pm_num, 'Point')
        ET.SubElement(pt_num, 'altitudeMode').text = 'clampToGround'
        ET.SubElement(pt_num, 'coordinates').text = f'{cen_lon},{cen_lat},0'

        # Site info dot — colored triangle, label hidden; tap to see full name + notes
        pm_lbl = ET.SubElement(folder, 'Placemark')
        ET.SubElement(pm_lbl, 'name').text = site
        ET.SubElement(pm_lbl, 'description').text = desc
        ET.SubElement(pm_lbl, 'styleUrl').text = f'#site{ci}'
        pt = ET.SubElement(pm_lbl, 'Point')
        ET.SubElement(pt, 'altitudeMode').text = 'clampToGround'
        ET.SubElement(pt, 'coordinates').text = f'{cen_lon},{cen_lat},0'

        # Track sub-folder — hidden at regional zoom, visible when zoomed into site
        track_folder = ET.SubElement(folder, 'Folder')
        ET.SubElement(track_folder, 'name').text = 'Tracks'
        lats = [c[0] for c in all_coords]
        lons = [c[1] for c in all_coords]
        pad = 0.1
        rgn = ET.SubElement(track_folder, 'Region')
        box = ET.SubElement(rgn, 'LatLonAltBox')
        ET.SubElement(box, 'north').text = f'{max(lats) + pad:.4f}'
        ET.SubElement(box, 'south').text = f'{min(lats) - pad:.4f}'
        ET.SubElement(box, 'east').text  = f'{max(lons) + pad:.4f}'
        ET.SubElement(box, 'west').text  = f'{min(lons) - pad:.4f}'
        lod = ET.SubElement(rgn, 'Lod')
        ET.SubElement(lod, 'minLodPixels').text = '128'
        ET.SubElement(lod, 'maxLodPixels').text = '-1'

        # Pass tracks
        for pi, p in enumerate(site_passes, 1):
            coords = p['coords']

            # Track line — visual only, no popup info needed
            pm_line = ET.SubElement(track_folder, 'Placemark')
            ET.SubElement(pm_line, 'name').text = ''
            ET.SubElement(pm_line, 'styleUrl').text = f'#line{ci}'
            ls_el = ET.SubElement(pm_line, 'LineString')
            ET.SubElement(ls_el, 'tessellate').text = '1'
            ET.SubElement(ls_el, 'altitudeMode').text = 'clampToGround'
            ET.SubElement(ls_el, 'coordinates').text = '\n'.join(
                f'{lon},{lat},0' for lat, lon in coords
            )

            if len(coords) < 2:
                continue

            # Start dot — green visual indicator only, no name/description
            pm_s = ET.SubElement(track_folder, 'Placemark')
            ET.SubElement(pm_s, 'name').text = ''
            ET.SubElement(pm_s, 'styleUrl').text = '#startDot'
            pt_s = ET.SubElement(pm_s, 'Point')
            ET.SubElement(pt_s, 'altitudeMode').text = 'clampToGround'
            ET.SubElement(pt_s, 'coordinates').text = f'{coords[0][1]},{coords[0][0]},0'

            # End dot — colored visual indicator only
            pm_e = ET.SubElement(track_folder, 'Placemark')
            ET.SubElement(pm_e, 'name').text = ''
            end_style = ET.SubElement(pm_e, 'Style')
            end_ic = ET.SubElement(end_style, 'IconStyle')
            ET.SubElement(end_ic, 'color').text = kc
            ET.SubElement(end_ic, 'scale').text = '0.65'
            end_ico = ET.SubElement(end_ic, 'Icon')
            ET.SubElement(end_ico, 'href').text = 'http://maps.google.com/mapfiles/kml/shapes/shaded_dot.png'
            ET.SubElement(ET.SubElement(end_style, 'LabelStyle'), 'scale').text = '0'
            pt_e = ET.SubElement(pm_e, 'Point')
            ET.SubElement(pt_e, 'altitudeMode').text = 'clampToGround'
            ET.SubElement(pt_e, 'coordinates').text = f'{coords[-1][1]},{coords[-1][0]},0'

    return kml

# ── Serialise with pretty-print ───────────────────────────────────────────────

def save_kml(kml_element, outfile):
    raw = ET.tostring(kml_element, encoding='unicode')
    def wrap_cdata(m):
        inner = html_mod.unescape(m.group(1))
        return f'<description><![CDATA[{inner}]]></description>'
    raw = re.sub(r'<description>(.*?)</description>', wrap_cdata, raw, flags=re.DOTALL)
    pretty = minidom.parseString(raw).toprettyxml(indent='  ')
    clean = '<?xml version="1.0" encoding="UTF-8"?>\n' + '\n'.join(pretty.split('\n')[1:])
    with open(outfile, 'w', encoding='utf-8') as f:
        f.write(clean)
    print(f"Saved: {outfile}")

# ── Generate all four KML files ───────────────────────────────────────────────

save_kml(build_kml(passes_by_site_goa_2021, '2021', 'Gulf of Alaska',
                   log_notes=log_notes_goa_2021),
         '2021_flighttracks.kml')

save_kml(build_kml(passes_by_site_goa_2024, '2024', 'Gulf of Alaska',
                   log_notes=log_notes_goa_2024),
         '2024_flighttracks.kml')

save_kml(build_kml(passes_by_site_ali_2022, '2022', 'Aleutian Islands',
                   log_notes=log_notes_ali_2022),
         '2022_flighttracks.kml')

save_kml(build_kml(passes_by_site_ali_2023, '2023', 'Aleutian Islands',
                   log_notes=log_notes_ali_2023),
         '2023_flighttracks.kml')
